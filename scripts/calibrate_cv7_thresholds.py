"""
CV-7 delta-threshold calibration.

Measures real per-visit-pair deltas (size/border/color) on a bounded,
seeded sample of the staged UQ Longitudinal data, and reports the
distributions used to pick GROWTH_PCT_THRESHOLD,
COMPACTNESS_DELTA_THRESHOLD, and COLOR_DELTA_E_THRESHOLD in
src/temporal/delta.py -- following the same "calibrate once against
real data, document it" discipline as
scripts/calibrate_cv1_resolution.py and
scripts/calibrate_cv6_temperature.py.

This script does NOT modify src/temporal/delta.py. It reports
distributions; the thresholds are set by hand afterward, informed by
this output, the same separation used in every prior calibration
script this project.

Bounded by design: caps the number of visit pairs processed (default
300) rather than running over the full staged sample, consistent with
every other bounded experiment in this project. One pair per lesion
(the first consecutive visit pair) is sampled, so coverage spreads
across many distinct lesions rather than concentrating on a few with
many visits.
"""

from __future__ import annotations

import argparse
import io
import random
import re
import zipfile
from collections import defaultdict
from pathlib import Path

import cv2
import numpy as np
import torch
from openpyxl import load_workbook

from src.segmentation.inference import load_segmentation_model, predict_mask
from src.temporal.calibration import calibrate
from src.temporal.delta import compute_delta
from src.temporal.measurement import measure_lesion

SEED = 11
DEFAULT_MAX_PAIRS = 300

REPO_ROOT = Path(__file__).resolve().parents[1]
ZIP_PATH = REPO_ROOT / "data/raw/UQ_zip/866990d01449152d_NIMARE-A11453_A11453.zip"
ARCHIVE_ROOT = "866990d01449152d_NIMARE-A11453_A11453/data/"
DERM_PREFIX = ARCHIVE_ROOT + "Dermoscopic Images/"
CHECKPOINT = REPO_ROOT / "checkpoints/cv3_512/best.pt"

OUTPUT_DIR = REPO_ROOT / "analysis/quality/cv7_temporal_data"

STAGED_PARTICIPANTS_FILES = [
    OUTPUT_DIR / "sampled_participants.txt",
    OUTPUT_DIR / "malignant_enrichment_participants.txt",
]

FILENAME_RE = re.compile(r"^(General|HighRisk)(\d+)_Lesion(\d+)_visit(\d+)(?:-\d+)?\.jpe?g$", re.IGNORECASE)

MALIGNANT_DIAGNOSES = {"melanoma", "basal cell carcinoma", "squamous cell carcinoma", "bcc", "scc"}


def parse_args():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--max-pairs", type=int, default=DEFAULT_MAX_PAIRS)
    return parser.parse_args()


def load_staged_participants() -> set[str]:
    participants = set()
    for path in STAGED_PARTICIPANTS_FILES:
        participants.update(line.strip() for line in path.read_text().splitlines() if line.strip())
    return participants


def load_diagnosis_lookup(zf: zipfile.ZipFile) -> dict[str, str]:
    """Map '{Cohort}{Participant}_Lesion{N}' -> lowercased Diagnosis string."""
    lookup: dict[str, str] = {}
    for member in (
        ARCHIVE_ROOT + "Dermoscopic Images/General Dermosopic images.xlsx",
        ARCHIVE_ROOT + "Dermoscopic Images/HighRisk Dermoscopic images.xlsx",
    ):
        with zf.open(member) as f:
            data = io.BytesIO(f.read())
        wb = load_workbook(data, read_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        next(rows)  # header
        for row in rows:
            if not row or not row[0]:
                continue
            match = FILENAME_RE.match(str(row[0]).strip())
            if not match:
                continue
            cohort, participant, lesion, _visit = match.groups()
            key = f"{cohort}{participant}_Lesion{lesion}"
            diagnosis = str(row[3]).strip().lower() if row[3] else "unknown"
            lookup[key] = diagnosis
    return lookup


def find_visit_pairs(zf: zipfile.ZipFile, staged_participants: set[str]) -> list[tuple[str, str, str]]:
    """
    Return (lesion_key, earlier_path, later_path) for the first
    consecutive visit pair of every staged lesion with >=2 visits.
    """
    names = [n for n in zf.namelist() if n.startswith(DERM_PREFIX) and n.lower().endswith((".jpg", ".jpeg"))]

    by_lesion: dict[str, list[tuple[int, str]]] = defaultdict(list)
    for name in names:
        basename = name.rsplit("/", 1)[-1]
        match = FILENAME_RE.match(basename)
        if not match:
            continue
        cohort, participant, lesion, visit = match.groups()
        if f"{cohort}{participant}" not in staged_participants:
            continue
        key = f"{cohort}{participant}_Lesion{lesion}"
        by_lesion[key].append((int(visit), name))

    pairs = []
    for key, visits in by_lesion.items():
        visits.sort(key=lambda v: v[0])
        if len(visits) >= 2:
            pairs.append((key, visits[0][1], visits[1][1]))
    return pairs


def _measure(zf: zipfile.ZipFile, path: str, model, device, cache: dict):
    if path in cache:
        return cache[path]
    with zf.open(path) as f:
        data = np.frombuffer(f.read(), dtype=np.uint8)
    image = cv2.imdecode(data, cv2.IMREAD_COLOR)
    calib = calibrate(image)
    resized = cv2.resize(image, (512, 512), interpolation=cv2.INTER_LINEAR)
    rgb = cv2.cvtColor(resized, cv2.COLOR_BGR2RGB)
    tensor = torch.from_numpy(rgb.astype(np.float32).transpose(2, 0, 1) / 255.0).unsqueeze(0)
    mask = predict_mask(model, tensor, device)
    result = measure_lesion(image, mask, calib)
    cache[path] = result
    return result


def main():
    args = parse_args()
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    staged_participants = load_staged_participants()
    device = torch.device("cpu")
    model = load_segmentation_model(CHECKPOINT, device)

    with zipfile.ZipFile(ZIP_PATH) as zf:
        diagnosis_lookup = load_diagnosis_lookup(zf)
        all_pairs = find_visit_pairs(zf, staged_participants)

        random.seed(SEED)
        random.shuffle(all_pairs)
        sampled_pairs = all_pairs[: args.max_pairs]

        cache: dict = {}
        rows = []
        for key, earlier_path, later_path in sampled_pairs:
            earlier = _measure(zf, earlier_path, model, device, cache)
            later = _measure(zf, later_path, model, device, cache)
            delta = compute_delta(earlier, later)
            diagnosis = diagnosis_lookup.get(key, "unknown")
            malignant = diagnosis in MALIGNANT_DIAGNOSES
            rows.append(
                {
                    "lesion": key,
                    "diagnosis": diagnosis,
                    "malignant": malignant,
                    "size_pct_change": delta.size_pct_change,
                    "border_delta": delta.border_delta,
                    "color_delta": delta.color_delta,
                    "confidence": delta.confidence,
                }
            )

    total = len(rows)
    both_masks_valid = sum(1 for r in rows if r["border_delta"] is not None)
    both_calibrated = sum(1 for r in rows if r["size_pct_change"] is not None)

    def percentiles(values, ps=(50, 75, 90, 95)):
        if not values:
            return {}
        arr = np.array(values)
        return {p: float(np.percentile(arr, p)) for p in ps}

    border_deltas = [abs(r["border_delta"]) for r in rows if r["border_delta"] is not None]
    color_deltas = [r["color_delta"] for r in rows if r["color_delta"] is not None]
    size_pct_changes = [abs(r["size_pct_change"]) for r in rows if r["size_pct_change"] is not None]

    malignant_rows = [r for r in rows if r["malignant"]]
    benign_rows = [r for r in rows if not r["malignant"]]

    report = []
    report.append(f"n pairs sampled: {total} (seed={SEED}, max_pairs={args.max_pairs})")
    report.append(f"both masks valid (border/color computable): {both_masks_valid} ({100*both_masks_valid/total:.1f}%)")
    report.append(f"both calibrated (size computable): {both_calibrated} ({100*both_calibrated/total:.1f}%)")
    report.append(f"malignant-outcome lesions in sample: {len(malignant_rows)}")
    report.append("")
    report.append(f"abs(border_delta) percentiles [n={len(border_deltas)}]: {percentiles(border_deltas)}")
    report.append(f"color_delta (Lab CIE76) percentiles [n={len(color_deltas)}]: {percentiles(color_deltas)}")
    report.append(f"abs(size_pct_change) percentiles [n={len(size_pct_changes)}]: {percentiles(size_pct_changes)}")
    report.append("")

    def mean(xs):
        return float(np.mean(xs)) if xs else None

    report.append(
        f"border_delta mean -- malignant: {mean([abs(r['border_delta']) for r in malignant_rows if r['border_delta'] is not None])}, "
        f"benign: {mean([abs(r['border_delta']) for r in benign_rows if r['border_delta'] is not None])}"
    )
    report.append(
        f"color_delta mean -- malignant: {mean([r['color_delta'] for r in malignant_rows if r['color_delta'] is not None])}, "
        f"benign: {mean([r['color_delta'] for r in benign_rows if r['color_delta'] is not None])}"
    )

    text = "\n".join(report)
    print(text)
    (OUTPUT_DIR / "delta_calibration_raw.txt").write_text(text + "\n")


if __name__ == "__main__":
    main()
